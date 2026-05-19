"""
Build SFW_chatgpt_ads_campaigns.xlsx by populating the OpenAI campaign workbook
template with SFW Construction's 10 active microsite services.

Source template: campaign_workbook_template.xlsx (downloaded from OpenAI Ads Manager Beta)
Output: SFW_chatgpt_ads_campaigns.xlsx + matching CSVs

Schema (from template):
  Sheet 'campaigns': campaign_name, budget_max, budget_type, launch_date, end_date,
                     objective, target_countries
  Sheet 'adgroups':  campaign_name, adgroup_name, max_bid, keywords, negative_keywords
  Sheet 'ads':       adgroup_name, title (<=24 chars), copy (<=48 chars), link, image_link

Header rows 1-4 are template metadata and MUST be preserved; data is appended row 5+.
"""
from __future__ import annotations
import csv
import json
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

HERE = Path(__file__).parent
TEMPLATE = HERE.parent / "campaign_workbook_template.xlsx"
OUT_XLSX = HERE / "SFW_chatgpt_ads_campaigns.xlsx"
OUT_CAMPAIGNS_CSV = HERE / "campaigns.csv"
OUT_ADGROUPS_CSV = HERE / "adgroups.csv"
OUT_ADS_CSV = HERE / "ads.csv"

LAUNCH_DATE = "2026-05-18"
END_DATE = "2026-07-15"
COUNTRIES = ["US"]  # ChatGPT Ads only supports US delivery today (per OpenAI docs).
UTM_TEMPLATE = (
    "utm_source=chatgpt&utm_medium=ads&utm_campaign={camp}&utm_content={adgroup}__{variant}"
)
# Brand-image fallback used when an ad group has no matching job photo.
# Square PNG/JPG, 640×640 minimum, 1200×1200 maximum, publicly accessible.
SFW_AD_IMAGE = "https://sfwconstruction.com/wp-content/uploads/2026/05/sfw-chatgpt-ad-1200.png"

# Per-ad-group 1200x1200 job photos (square JPG, <500KB, public).
# Hosted in gs://buildertrend-rip/sfw-ad-images/. Picks come from
# image_matches_from_postman.md (primary match per ad group).
_GCS = "https://storage.googleapis.com/buildertrend-rip/sfw-ad-images"
AD_GROUP_IMAGE_URLS: dict[str, str] = {
    "SFW_DeckRepair_RotDamage":  f"{_GCS}/DeckRepair/post62_509762196-IMG_8181.jpg",
    "SFW_DeckRepair_Structural": f"{_GCS}/DeckRepair/post7_368403905-20210927_130441.jpg",
    "SFW_DeckRepair_Inspection": f"{_GCS}/DeckRepair/post21_782562654-PXL_20240613_210543640.jpg",
    "SFW_Chimney_WoodRot":       f"{_GCS}/Chimney/post29_733622966-IMG_1162.jpg",
    "SFW_Chimney_Leaking":       f"{_GCS}/Chimney/post29_733622966-IMG_1162.jpg",
    "SFW_Chimney_Inspection":    f"{_GCS}/Chimney/post29_733622966-IMG_1162.jpg",
    "SFW_Siding_HardieBoard":    f"{_GCS}/Siding-HardieBoard/post6_1039530952-PXL_20250923_004243722.jpg",
    "SFW_Siding_RotRepair":      f"{_GCS}/Siding-RotRepair/post24_495728415-IMG_1801.jpg",
    "SFW_Siding_Replacement":    f"{_GCS}/Siding-Replacement/post16_330957302-PXL_20210604_192939386.jpg",
    "SFW_Crawlspace_RotRepair":  f"{_GCS}/Crawlspace/post69_918374393-20250311_144451.jpg",
    "SFW_Crawlspace_Subfloor":   f"{_GCS}/Crawlspace/post69_918374393-20250311_144451.jpg",
    "SFW_Crawlspace_Moisture":   f"{_GCS}/Crawlspace/post69_918374393-20250311_144451.jpg",
    "SFW_Leak_WindowLeak":       f"{_GCS}/Leak-Window/post11_325099932-20210521_115753.jpg",
    "SFW_Leak_WaterIntrusion":   f"{_GCS}/Leak-WaterIntrusion/post67_922368358-20250318_114421.jpg",
    "SFW_Flashing_RoofWall":     f"{_GCS}/Flashing/post4_1067070769-IMG_6577.jpg",
    "SFW_Flashing_WindowKickout":f"{_GCS}/Flashing/post86_953225577-PXL_20250506_172943044.jpg",
    "SFW_LeadPaint_Removal":     f"{_GCS}/LeadPaint/post18_377872714-IMG_0302.jpg",
    "SFW_LeadPaint_Testing":     f"{_GCS}/LeadPaint/post18_377872714-IMG_0302.jpg",
    "SFW_DryRot_Repair":         f"{_GCS}/DryRot/post63_917631036-20250310_112625.jpg",
    "SFW_DryRot_Prevention":     f"{_GCS}/DryRot/post67_922368358-20250318_114421.jpg",
    "SFW_Trim_FasciaSoffit":     f"{_GCS}/Trim-FasciaSoffit/post10_835213092-PXL_20240926_211850068.jpg",
    "SFW_Trim_Exterior":         f"{_GCS}/Trim-Exterior/post5_1061573684-IMG_6431.jpg",
    "SFW_Beam_Structural":       f"{_GCS}/Beam-Structural/post69_918374393-20250311_144451.jpg",
    "SFW_Beam_Sistering":        f"{_GCS}/Beam-Sistering/post68_900753867-20250206_144539.jpg",
}
# Every ad on every campaign lands on sfwconstruction.com — single advertiser brand.
SFW_DOMAIN = "https://sfwconstruction.com"


@dataclass
class Ad:
    title: str
    copy: str
    variant: str  # short slug, e.g. "problem-v1"

    def __post_init__(self) -> None:
        if len(self.title) > 24:
            raise ValueError(f"Title too long ({len(self.title)} chars): {self.title!r}")
        if len(self.copy) > 48:
            raise ValueError(f"Copy too long ({len(self.copy)} chars): {self.copy!r}")


@dataclass
class AdGroup:
    name: str  # unique across the whole account
    max_bid: float
    keywords: list[str]  # context hints
    negative_keywords: list[str] = field(default_factory=list)
    ads: list[Ad] = field(default_factory=list)

    def __post_init__(self) -> None:
        if len(self.negative_keywords) > 25:
            raise ValueError(
                f"Ad group {self.name!r}: {len(self.negative_keywords)} negative keywords "
                f"exceeds OpenAI's per-ad-group cap of 25"
            )


@dataclass
class Campaign:
    name: str
    service_slug: str
    landing_path: str  # path on sfwconstruction.com, e.g. "/repair-services/portland-deck-repair/"
    budget_max: float
    objective: str  # "Clicks" or "Views"
    ad_groups: list[AdGroup] = field(default_factory=list)
    budget_type: str = "Lifetime"

    @property
    def landing_url(self) -> str:
        return SFW_DOMAIN + self.landing_path


# ---------------------------------------------------------------------------
# Campaigns
# ---------------------------------------------------------------------------

CAMPAIGNS: list[Campaign] = [
    # ---- Tier 1: high-value, competitive --------------------------------
    Campaign(
        name="SFW_DeckRepair_PDX_2026Q2",
        service_slug="deck-repair",
        landing_path="/repair-services/portland-deck-repair/",
        budget_max=1000,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_DeckRepair_RotDamage",
                max_bid=4.0,
                keywords=[
                    "homeowner in Portland Oregon asking how to repair a rotted wood deck",
                    "signs of dry rot in deck boards joists or posts",
                    "Pacific Northwest deck rot from rain and moisture",
                    "soft or spongy deck boards that need replacement",
                    "rotted deck repair vs full deck replacement cost",
                    "deck rot inspection Portland metro",
                ],
                negative_keywords=["diy", "youtube tutorial", "home depot", "lowes"],
                ads=[
                    Ad("Portland Deck Rot Repair", "Sagging, soft, or rotted? Free inspection.", "problem-v1"),
                    Ad("Deck Rot Repair Experts", "Licensed Portland deck pros. Free estimate.", "trust-v1"),
                    Ad("Fix Your Rotted Deck", "Replace rotted boards & joists. Portland, OR.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_DeckRepair_Structural",
                max_bid=4.0,
                keywords=[
                    "wobbly or unsafe deck structural repair",
                    "sagging deck joists and beams needing replacement",
                    "load-bearing deck post and beam restoration",
                    "Portland homeowner asking if a deck is safe to walk on",
                    "deck collapse risk assessment and structural repair",
                ],
                negative_keywords=["diy", "deck building plans"],
                ads=[
                    Ad("Deck Structural Repair", "Wobbly, unsafe deck? We rebuild it right.", "problem-v1"),
                    Ad("Portland Deck Builders", "Beams, joists & decking. Licensed in OR.", "trust-v1"),
                    Ad("Save Your Deck", "Repair, don't replace. Free PDX inspection.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_DeckRepair_Inspection",
                max_bid=4.0,
                keywords=[
                    "free deck inspection in Portland Oregon",
                    "what to look for in a deck inspection before buying a home",
                    "deck contractor estimate Portland metro",
                    "how much does deck repair cost in Portland",
                ],
                negative_keywords=["home inspection app", "diy inspection checklist"],
                ads=[
                    Ad("Free Deck Inspection", "Portland licensed contractor. Same-week visit.", "offer-v1"),
                    Ad("Deck Estimate Portland", "Honest pricing. Photos & written estimate.", "trust-v1"),
                ],
            ),
        ],
    ),
    Campaign(
        name="SFW_ChimneyRepair_PDX_2026Q2",
        service_slug="chimney-repair",
        landing_path="/chimney-chase-repair/",
        budget_max=1000,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_Chimney_WoodRot",
                max_bid=5.0,
                keywords=[
                    "wood-framed chimney rot repair in Pacific Northwest",
                    "rotted siding and framing around a wood chimney chase",
                    "homeowner asking how to fix a rotted chimney chase",
                    "Portland chimney framing replacement after water damage",
                    "chimney chase rot causing roof leaks",
                ],
                negative_keywords=["masonry chimney", "brick chimney repair", "diy"],
                ads=[
                    Ad("Wood Chimney Rot Repair", "Restore framing, siding & flashing. PDX.", "problem-v1"),
                    Ad("Chimney Rot Specialists", "Stop the rot. Portland-licensed crew.", "trust-v1"),
                    Ad("Rotted Chimney? Fix It.", "Full restoration & rebuild. Free estimate.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Chimney_Leaking",
                max_bid=5.0,
                keywords=[
                    "chimney leak around flashing or chase top",
                    "water stains on ceiling near chimney Portland",
                    "how to stop a leaking chimney in a wet climate",
                    "chimney flashing failure causing water intrusion",
                ],
                negative_keywords=["chimney sweep", "fireplace cleaning", "diy"],
                ads=[
                    Ad("Chimney Leak Repair", "Seal flashing & stop water damage.", "problem-v1"),
                    Ad("Portland Chimney Pros", "Leak detection & repair. Same-week.", "trust-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Chimney_Inspection",
                max_bid=4.0,
                keywords=[
                    "free chimney inspection in Portland Oregon",
                    "chimney rot inspection before selling a home",
                    "chimney structural assessment Pacific Northwest",
                ],
                negative_keywords=["chimney sweep cost", "creosote removal"],
                ads=[
                    Ad("Free Chimney Inspection", "Portland-area experts. No-pressure quote.", "offer-v1"),
                    Ad("Chimney Repair Experts", "30+ years framing wood chimneys in PDX.", "trust-v1"),
                ],
            ),
        ],
    ),
    Campaign(
        name="SFW_SidingRepair_PDX_2026Q2",
        service_slug="siding-repair",
        landing_path="/siding-repair-portland/",
        budget_max=1000,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_Siding_HardieBoard",
                max_bid=5.0,
                keywords=[
                    "Hardie board siding installation in Portland Oregon",
                    "James Hardie fiber cement siding repair specialists",
                    "matching and patching damaged Hardie plank siding",
                    "Hardie board contractor Pacific Northwest",
                ],
                negative_keywords=["diy hardie board", "hardie siding cost calculator"],
                ads=[
                    Ad("Hardie Board Experts", "Install & repair. Portland licensed crew.", "trust-v1"),
                    Ad("Hardie Siding Repair", "Match, patch, or replace. Free estimate.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Siding_RotRepair",
                max_bid=5.0,
                keywords=[
                    "rotted siding repair Portland metro",
                    "water-damaged exterior siding replacement",
                    "Pacific Northwest siding rot from rain exposure",
                    "homeowner asking how to fix soft or peeling siding",
                ],
                negative_keywords=["diy siding repair", "siding paint"],
                ads=[
                    Ad("Rotted Siding Repair", "Replace damaged boards. Portland, OR.", "problem-v1"),
                    Ad("Stop Siding Rot", "Fix water-damaged siding. Free inspection.", "solution-v1"),
                    Ad("Siding Rot Specialists", "Pacific NW moisture experts. Licensed.", "trust-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Siding_Replacement",
                max_bid=5.0,
                keywords=[
                    "full siding replacement Portland Oregon homeowner",
                    "cedar siding replacement Pacific Northwest",
                    "exterior siding remodel contractor estimate",
                    "T1-11 plywood siding replacement",
                ],
                negative_keywords=["vinyl siding wholesale", "siding manufacturer"],
                ads=[
                    Ad("Portland Siding Pros", "Full or partial replacement. Free estimate.", "trust-v1"),
                    Ad("Replace Old Siding", "Hardie & cedar specialists. PDX licensed.", "solution-v1"),
                ],
            ),
        ],
    ),
    Campaign(
        name="SFW_CrawlspaceRot_PDX_2026Q2",
        service_slug="crawlspace-rot",
        landing_path="/repair-services/crawl-space-repair-portland/",
        budget_max=1000,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_Crawlspace_RotRepair",
                max_bid=5.0,
                keywords=[
                    "crawl space rot repair Portland Oregon",
                    "rotted subfloor and floor joists under a Pacific Northwest home",
                    "homeowner asking about wood rot under the house",
                    "sill plate and rim joist rot replacement",
                    "musty smell from crawl space and possible rot",
                ],
                negative_keywords=["crawl space cleaning service", "pest control", "diy"],
                ads=[
                    Ad("Crawl Space Rot Repair", "Subfloor & joist replacement. Portland, OR.", "problem-v1"),
                    Ad("Rotted Subfloor Repair", "Restore strength under your home.", "solution-v1"),
                    Ad("Stop Crawl Space Rot", "Repair & moisture control. Free inspection.", "offer-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Crawlspace_Subfloor",
                max_bid=5.0,
                keywords=[
                    "sagging floor caused by rotted joists Portland",
                    "subfloor replacement under bathroom or kitchen",
                    "sister joist repair to restore floor strength",
                    "bouncy floor diagnosis Pacific Northwest home",
                ],
                negative_keywords=["squeaky floor diy", "subfloor adhesive"],
                ads=[
                    Ad("Sagging Floor Repair", "Sister or replace damaged joists. PDX.", "problem-v1"),
                    Ad("Subfloor Repair PDX", "Replace rotted sheathing & supports.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Crawlspace_Moisture",
                max_bid=4.0,
                keywords=[
                    "crawl space moisture control Portland Oregon",
                    "vapor barrier and encapsulation Pacific Northwest",
                    "wet crawl space causing wood rot",
                    "crawl space drainage and ventilation contractor",
                ],
                negative_keywords=["crawl space dehumidifier brand", "diy encapsulation kit"],
                ads=[
                    Ad("Crawl Space Moisture", "Vapor barriers & encapsulation. Portland.", "solution-v1"),
                    Ad("Wet Crawl Space?", "Diagnose & repair. Free PDX inspection.", "problem-v1"),
                ],
            ),
        ],
    ),
    # ---- Tier 2: emergency / urgent intent ------------------------------
    Campaign(
        name="SFW_LeakRepair_PDX_2026Q2",
        service_slug="leak-repair",
        landing_path="/locations/portland/portland-window-leak-repair/",
        budget_max=600,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_Leak_WindowLeak",
                max_bid=5.0,
                keywords=[
                    "leaking window repair Portland Oregon",
                    "water dripping inside around a window during rain",
                    "window flashing failure causing wall leak",
                    "homeowner asking how to stop a window from leaking",
                ],
                negative_keywords=["window replacement quote", "diy caulk window"],
                ads=[
                    Ad("Window Leak Repair", "Stop leaks before damage spreads. PDX.", "problem-v1"),
                    Ad("Leaking Window? Fix It", "Reseal, reflash, restore. Free estimate.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Leak_WaterIntrusion",
                max_bid=5.0,
                keywords=[
                    "hidden water intrusion in exterior walls Portland",
                    "finding the source of a mystery leak in a home",
                    "Pacific Northwest home water damage from rain",
                    "moisture meter leak detection contractor",
                ],
                negative_keywords=["plumber", "burst pipe", "water heater leak"],
                ads=[
                    Ad("Water Intrusion Pros", "Find & seal hidden leaks. Portland metro.", "trust-v1"),
                    Ad("Hidden Leak Detection", "Pinpoint the source. Repair & restore.", "solution-v1"),
                ],
            ),
        ],
    ),
    Campaign(
        name="SFW_FlashingRepair_PDX_2026Q2",
        service_slug="flashing-repair",
        landing_path="/repair-services/roof-repair-flat-roof-repair/",
        budget_max=600,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_Flashing_RoofWall",
                max_bid=4.0,
                keywords=[
                    "roof flashing repair Portland Oregon",
                    "chimney flashing leak fix",
                    "valley flashing replacement Pacific Northwest",
                    "wall flashing failure causing siding rot",
                ],
                negative_keywords=["full roof replacement", "diy roof flashing", "roofing materials"],
                ads=[
                    Ad("Flashing Repair Pros", "Stop leaks at the source. Portland, OR.", "trust-v1"),
                    Ad("Roof Flashing Repair", "Chimney, valley & wall flashing. Licensed.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Flashing_WindowKickout",
                max_bid=4.0,
                keywords=[
                    "kickout flashing installation Portland",
                    "window flashing repair to stop water intrusion",
                    "missing kickout diverter causing hidden wall rot",
                ],
                negative_keywords=["window replacement", "diy flashing tape"],
                ads=[
                    Ad("Window Flashing Repair", "Prevent leaks & rot. Portland metro.", "problem-v1"),
                    Ad("Kickout Flashing Pros", "Stop hidden wall rot. Free inspection.", "offer-v1"),
                ],
            ),
        ],
    ),
    # ---- Tier 3: specialty / lower-volume -------------------------------
    Campaign(
        name="SFW_LeadPaint_PDX_2026Q2",
        service_slug="lead-paint",
        landing_path="/locations/portland/house-painting-portland/lead-based-paint-removal/",
        budget_max=400,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_LeadPaint_Removal",
                max_bid=3.0,
                keywords=[
                    "lead paint removal in a pre-1978 Portland home",
                    "EPA RRP-certified lead-safe renovation contractor",
                    "lead paint abatement and encapsulation Pacific Northwest",
                    "homeowner asking about safely removing old lead paint",
                ],
                negative_keywords=["diy lead paint removal", "lead paint cost calculator"],
                ads=[
                    Ad("Lead Paint Removal", "EPA-certified. Portland licensed crew.", "trust-v1"),
                    Ad("Lead-Safe Renovation", "RRP-certified contractor. Free consult.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_LeadPaint_Testing",
                max_bid=3.0,
                keywords=[
                    "lead paint testing for an older Portland home",
                    "XRF lead testing before remodel or renovation",
                    "lab analysis lead paint inspection Pacific Northwest",
                ],
                negative_keywords=["lead test kit amazon", "diy lead swab"],
                ads=[
                    Ad("Lead Paint Testing", "Certified XRF & lab analysis. Portland.", "solution-v1"),
                    Ad("Pre-1978 Home? Test.", "Protect your family. Certified testing.", "problem-v1"),
                ],
            ),
        ],
    ),
    Campaign(
        name="SFW_DryRot_PDX_2026Q2",
        service_slug="dry-rot",
        landing_path="/repair-services/dry-rot-repair/",
        budget_max=400,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_DryRot_Repair",
                max_bid=4.0,
                keywords=[
                    "dry rot repair Portland Oregon home",
                    "removing and replacing rotted wood framing",
                    "wood-decay fungus identification and treatment Pacific Northwest",
                    "homeowner asking about soft or punky wood on the exterior",
                ],
                negative_keywords=["diy wood filler", "rot epoxy product brand"],
                ads=[
                    Ad("Dry Rot Repair Pros", "Remove & replace rotted wood. Portland.", "trust-v1"),
                    Ad("Stop Wood Rot", "Find the source, fix it right. Licensed.", "solution-v1"),
                    Ad("Dry Rot Specialists", "Pacific NW moisture & rot experts.", "trust-v2"),
                ],
            ),
            AdGroup(
                name="SFW_DryRot_Prevention",
                max_bid=3.0,
                keywords=[
                    "preventing wood rot in a Pacific Northwest climate",
                    "treated lumber and moisture barrier strategy",
                    "borate wood treatment for rot prevention",
                ],
                negative_keywords=["wood preservative brand", "diy borate"],
                ads=[
                    Ad("Wood Rot Prevention", "Treat, seal & prevent recurrence. PDX.", "solution-v1"),
                ],
            ),
        ],
    ),
    Campaign(
        name="SFW_TrimRepair_PDX_2026Q2",
        service_slug="trim-repair",
        # No dedicated trim page on sfwconstruction.com; rot repair is the closest intent
        # match (rotted trim is the dominant search intent for this service).
        landing_path="/repair-services/dry-rot-repair/",
        budget_max=400,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_Trim_FasciaSoffit",
                max_bid=3.0,
                keywords=[
                    "rotted fascia board repair Portland Oregon",
                    "soffit replacement to restore attic ventilation",
                    "Pacific Northwest fascia and soffit rot from moisture",
                ],
                negative_keywords=["gutter installation", "diy fascia"],
                ads=[
                    Ad("Fascia & Soffit Repair", "Replace rotted boards. Portland licensed.", "solution-v1"),
                    Ad("Soffit Replacement", "Restore curb appeal & ventilation. PDX.", "trust-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Trim_Exterior",
                max_bid=3.0,
                keywords=[
                    "exterior window and door trim repair Portland",
                    "corner board and rake trim replacement",
                    "rotted exterior trim restoration Pacific Northwest",
                ],
                negative_keywords=["interior trim", "molding catalog"],
                ads=[
                    Ad("Exterior Trim Repair", "Window, door & corner board restore.", "solution-v1"),
                    Ad("Rotted Trim Repair", "Match & replace damaged trim. Free quote.", "problem-v1"),
                ],
            ),
        ],
    ),
    Campaign(
        name="SFW_BeamRepair_PDX_2026Q2",
        service_slug="beam-repair",
        # Beam/structural work lives under the construction-defect repair service page.
        landing_path="/repair-services/construction-defect-repair-portland/",
        budget_max=400,
        objective="Clicks",
        ad_groups=[
            AdGroup(
                name="SFW_Beam_Structural",
                max_bid=4.0,
                keywords=[
                    "structural beam repair Portland Oregon",
                    "load-bearing beam replacement in an older home",
                    "engineered lumber beam restoration Pacific Northwest",
                    "homeowner asking if a damaged beam can be repaired",
                ],
                negative_keywords=["lumber pricing", "diy structural repair"],
                ads=[
                    Ad("Beam Repair Experts", "Load-bearing beam restoration. PDX.", "trust-v1"),
                    Ad("Load-Bearing Beam Fix", "Sister or replace. Engineered solutions.", "solution-v1"),
                ],
            ),
            AdGroup(
                name="SFW_Beam_Sistering",
                max_bid=4.0,
                keywords=[
                    "beam sistering to reinforce weakened framing",
                    "sagging beam repair without full replacement",
                    "structural reinforcement Portland Oregon contractor",
                ],
                negative_keywords=["lvl beam calculator", "diy sistering"],
                ads=[
                    Ad("Beam Sistering Pros", "Reinforce without full replace. PDX.", "solution-v1"),
                    Ad("Sagging Beam Repair", "Restore structural integrity. Licensed.", "problem-v1"),
                ],
            ),
        ],
    ),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def ad_url(campaign: Campaign, ad_group: AdGroup, ad: Ad) -> str:
    utms = UTM_TEMPLATE.format(
        camp=campaign.service_slug,
        adgroup=slug_for_utm(ad_group.name),
        variant=ad.variant,
    )
    sep = "&" if "?" in campaign.landing_url else "?"
    return f"{campaign.landing_url}{sep}{utms}"


def slug_for_utm(name: str) -> str:
    return name.replace("SFW_", "").replace("_", "-").lower()


def image_url(campaign: Campaign, ad_group: AdGroup) -> str:
    del campaign
    return AD_GROUP_IMAGE_URLS.get(ad_group.name, SFW_AD_IMAGE)


def json_array(items: Iterable[str]) -> str:
    """OpenAI's bulk-upload format for keywords/countries is a JSON array literal."""
    return json.dumps(list(items))


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def rows_campaigns() -> list[list]:
    rows = []
    for c in CAMPAIGNS:
        rows.append([
            c.name,
            c.budget_max,
            c.budget_type,
            LAUNCH_DATE,
            END_DATE,
            c.objective,
            json_array(COUNTRIES),
        ])
    return rows


def rows_adgroups() -> list[list]:
    rows = []
    for c in CAMPAIGNS:
        for g in c.ad_groups:
            rows.append([
                c.name,
                g.name,
                g.max_bid,
                json_array(g.keywords),
                json_array(g.negative_keywords) if g.negative_keywords else "",
            ])
    return rows


def rows_ads() -> list[list]:
    rows = []
    for c in CAMPAIGNS:
        for g in c.ad_groups:
            for a in g.ads:
                rows.append([
                    g.name,
                    a.title,
                    a.copy,
                    ad_url(c, g, a),
                    image_url(c, g),
                ])
    return rows


def write_csv(path: Path, header: list[str], rows: list[list]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in rows:
            w.writerow(row)


def write_workbook() -> None:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE}")
    shutil.copyfile(TEMPLATE, OUT_XLSX)
    wb = load_workbook(OUT_XLSX)

    # Template rows 1-4 are metadata (column name, required/optional, description, example).
    # OpenAI's processor skips rows 1-4 — data begins on row 5.
    DATA_START_ROW = 5

    def fill(sheet_name: str, rows: list[list]) -> None:
        ws = wb[sheet_name]
        for r_idx, row in enumerate(rows, start=DATA_START_ROW):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    fill("campaigns", rows_campaigns())
    fill("adgroups", rows_adgroups())
    fill("ads", rows_ads())
    wb.save(OUT_XLSX)


def validate() -> list[str]:
    errors: list[str] = []
    campaign_names = [c.name for c in CAMPAIGNS]
    if len(set(campaign_names)) != len(campaign_names):
        errors.append("Duplicate campaign names")

    adgroup_names: list[str] = []
    for c in CAMPAIGNS:
        for g in c.ad_groups:
            adgroup_names.append(g.name)
            if c.objective == "Clicks" and g.max_bid is None:
                errors.append(f"Ad group {g.name!r}: Clicks campaign requires max_bid")
            if c.objective == "Views" and g.max_bid is not None:
                # not strictly an error, just informational
                pass
    if len(set(adgroup_names)) != len(adgroup_names):
        errors.append("Duplicate ad group names across the account")
    return errors


def main() -> int:
    errs = validate()
    if errs:
        print("Validation errors:", file=sys.stderr)
        for e in errs:
            print(" -", e, file=sys.stderr)
        return 1

    write_csv(
        OUT_CAMPAIGNS_CSV,
        ["campaign_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"],
        rows_campaigns(),
    )
    write_csv(
        OUT_ADGROUPS_CSV,
        ["campaign_name", "adgroup_name", "max_bid", "keywords", "negative_keywords"],
        rows_adgroups(),
    )
    write_csv(
        OUT_ADS_CSV,
        ["adgroup_name", "title", "copy", "link", "image_link"],
        rows_ads(),
    )

    write_workbook()

    n_camps = len(CAMPAIGNS)
    n_ags = sum(len(c.ad_groups) for c in CAMPAIGNS)
    n_ads = sum(len(g.ads) for c in CAMPAIGNS for g in c.ad_groups)
    total_budget = sum(c.budget_max for c in CAMPAIGNS)
    print(f"Wrote {OUT_XLSX.name}: {n_camps} campaigns, {n_ags} ad groups, {n_ads} ads")
    print(f"Total lifetime budget across all campaigns: ${total_budget:,.0f}")
    print(f"CSVs: {OUT_CAMPAIGNS_CSV.name}, {OUT_ADGROUPS_CSV.name}, {OUT_ADS_CSV.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
